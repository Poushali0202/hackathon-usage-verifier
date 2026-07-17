"""
Batch feeder for the hackathon usage verifier — HYBRID architecture.

Python gathers the GitHub evidence (complete, no truncation, fast); the RocketRide
pipeline (chat -> Claude -> answer) applies the rubric and writes the verdict.

MODES:
  1) Full run:  python run_batch.py "submissions.csv"
  2) Resubmit:  python run_batch.py "needs_review.csv" --merge RocketRide_Hackathon_Usage.xlsx

OPTIONS: --out FILE | --concurrency N | --limit N

For the full ~35-repo run, set ROCKETRIDE_GITHUB_TOKEN in .env to avoid GitHub's
60-requests/hour unauthenticated limit. Not needed for a small (<~6 repo) test.
"""

import argparse
import asyncio
import csv
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rocketride import RocketRideClient
from rocketride.schema import Question

PIPELINE_FILE = "verify_usage.pipe"
NEEDS_REVIEW = "needs_review.csv"
GH_TOKEN = None  # set in main()

# ---- column resolution (headers vary: spaces, "(ALL)", case) -----------------

FIELD_ALIASES = {
    "project": ["projecttitle", "projectname", "project", "name", "title"],
    "names": ["teammembersnamesall", "teammembersnames", "names", "teammembers", "teamname"],
    "emails": ["teammembersemailsall", "teammembersemails", "teamemails", "emails", "email"],
    "github": ["githubrepo", "githuburl", "githublink", "github", "gitlink", "gitrepo",
               "gitrepository", "gitrepolink", "repository", "repolink", "repositoryurl", "repo", "codelink"],
    "feedback": ["feedback"],
    "demo": ["videodemopresentation", "demopresentation", "demovideo", "demo", "presentation", "video"],
    "deployed": ["deployedprojecturl", "deployedurl", "liveappurl", "liveapp", "livelink", "liveurl",
                 "hostedurl", "deployment"],
}


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _resolve_headers(headers: list) -> dict:
    """Map CSV/XLSX headers to canonical fields. Two passes: exact alias match first
    (reliable), then substring — but substring only uses aliases >=5 chars so a short,
    generic alias like 'name' can't grab 'Team Members Names' or 'repo' grab 'Reporting'."""
    mapping = {}
    normed = [(h, _norm(h)) for h in headers]
    used = set()

    def claim(canonical, pred):
        for h, nh in normed:
            if h in used:
                continue
            if pred(nh):
                mapping[canonical] = h
                used.add(h)
                return

    for canonical, aliases in FIELD_ALIASES.items():          # pass 1: exact
        claim(canonical, lambda nh, al=aliases: nh in al)
    for canonical, aliases in FIELD_ALIASES.items():          # pass 2: substring (longer aliases only)
        if canonical not in mapping:
            longs = [a for a in aliases if len(a) >= 5]
            claim(canonical, lambda nh, al=longs: any(a in nh for a in al))
    return mapping


def load_rows(path: str) -> list:
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".csv":
        try:
            with open(p, encoding="utf-8-sig", newline="") as f:
                raw = [r for r in csv.reader(f) if any(c.strip() for c in r)]
        except UnicodeDecodeError:
            # some exports are Windows-encoded (cp1252), not UTF-8 — fall back
            with open(p, encoding="cp1252", newline="") as f:
                raw = [r for r in csv.reader(f) if any(c.strip() for c in r)]
    elif ext in (".xlsx", ".xlsm", ".xls"):
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        raw = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(c.strip() for c in cells):
                raw.append(cells)
    else:
        sys.exit(f"Unsupported file type: {ext} (use .csv or .xlsx)")
    if not raw:
        sys.exit("No rows found in the input file.")
    # Find the real header row: some exports put a banner/title row first (e.g. "LIVE TEAM RESULTS"),
    # so scan the first several rows for the one that actually resolves the required columns.
    # We need a GitHub link plus SOME label — a project column, or failing that a team/name column.
    header_row, mapping, headers = 0, {}, [h.strip() for h in raw[0]]
    for i in range(min(8, len(raw))):
        hs = [str(h).strip() for h in raw[i]]
        m = _resolve_headers(hs)
        if "github" in m and ("project" in m or "names" in m):
            header_row, mapping, headers = i, m, hs
            break
    if "github" not in mapping or ("project" not in mapping and "names" not in mapping):
        sys.exit("Could not find the required columns (need a GitHub link and a project or team name). "
                 f"Detected: {mapping}\nHeaders: {headers}")
    if "project" not in mapping:                 # no dedicated project column -> use the team/name column
        mapping["project"] = mapping["names"]
    idx = {c: headers.index(h) for c, h in mapping.items()}
    rows = [
        {c: (cells[i].strip() if i < len(cells) else "") for c, i in idx.items()}
        for cells in raw[header_row + 1:]
    ]
    for row in rows:  # a malformed multi-line cell can bleed following text into the title
        if "\n" in row.get("project", ""):
            row["project"] = row["project"].split("\n", 1)[0].strip()
    return rows


# ---- repo helpers ------------------------------------------------------------

MISSING = {"", "na", "n/a", "none", "nil", "-", "tbd"}


def repo_missing(url: str) -> bool:
    return _norm(url) in {_norm(m) for m in MISSING} or "github.com" not in url.lower()


def normalize_repo(url: str) -> str:
    m = re.search(r"github\.com[/:]+([^/\s]+/[^/\s#?]+)", url, re.I)
    return m.group(1).lower().removesuffix(".git") if m else url.lower().strip()


def extract_json(text: str) -> dict:
    if not text:
        return {}
    fence = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.S)
    candidate = fence.group(1) if fence else text
    brace = re.search(r"\{.*\}", candidate, re.S)
    if not brace:
        return {}
    try:
        return json.loads(brace.group(0))
    except json.JSONDecodeError:
        return {}


def is_failed(res: dict) -> bool:
    """Failed the access check = repo wasn't cleanly fetched/classified."""
    return res.get("repo_accessible") is not True or bool(res.get("classify_failed"))


# ---- GitHub evidence gathering (Python — reliable, no truncation) ------------

def github_token() -> str:
    t = os.environ.get("ROCKETRIDE_GITHUB_TOKEN")
    if t and not t.startswith("PASTE"):
        return t
    envf = Path(".env")
    if envf.exists():
        for line in envf.read_text(encoding="utf-8-sig").splitlines():
            if line.strip().startswith("ROCKETRIDE_GITHUB_TOKEN="):
                v = line.split("=", 1)[1].strip()
                if v and not v.startswith("PASTE"):
                    return v
    return None


def _gh(url: str, retries: int = 2) -> tuple:
    headers = {"User-Agent": "rr-verifier", "Accept": "application/vnd.github+json"}
    if GH_TOKEN:
        headers["Authorization"] = f"Bearer {GH_TOKEN}"
    for attempt in range(retries + 1):
        req = urllib.request.Request(url, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                return r.status, r.read().decode("utf-8", "replace")
        except urllib.error.HTTPError as e:
            if e.code in (403, 429, 500, 502, 503) and attempt < retries:
                time.sleep(1.5 * (attempt + 1))
                continue
            return e.code, ""
        except Exception:
            if attempt < retries:
                time.sleep(1.0)
                continue
            return None, ""
    return None, ""


def parse_repo(url: str):
    m = re.search(r"github\.com[/:]+([^/\s]+)/([^/\s#?]+)", url, re.I)
    return (m.group(1), m.group(2).removesuffix(".git")) if m else None


# Strong signals only — a bare "rocketride" mention (e.g. in a generic pipeline.js
# lookalike) is NOT counted; require real SDK/engine usage OR hosted-pipeline usage.
SRC_PATS = ["rocketrideclient", "client.use(", "client.chat(", "ws://localhost:5565",
            "/v1/pipelines", "@rocketride/sdk", "from rocketride", "import rocketride",
            "require('rocketride", 'require("rocketride',
            # hosted / Cloud usage: pipelines called by id via API + a local adapter module
            "rocketride_pipeline", "rocketride_api_key", "rocketride_apikey",
            "rocketride_uri", "lib/rocketride", "/rocketride'", '/rocketride"']

# Other major platforms — used to judge (from code, not feedback) whether RocketRide is the
# sole backbone (Yes) or sits alongside/beneath another platform (Partial).
OTHER_PLATFORMS = ["butterbase", "supabase", "xtrace", "photon", "langchain",
                   "crewai", "firebase", "pinecone", "weaviate"]


def fetch_signals(url: str) -> dict:
    """Return a compact, complete evidence digest for one repo (blocking I/O)."""
    pr = parse_repo(url)
    if not pr:
        return {"accessible": False, "note": "URL not parseable"}
    owner, repo = pr

    status, body = _gh(f"https://api.github.com/repos/{owner}/{repo}")
    if status != 200:
        return {"accessible": False, "status": status}
    branch = json.loads(body).get("default_branch", "main")

    st, tbody = _gh(f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}?recursive=1")
    if st != 200:
        # repo is reachable but we couldn't read its file tree — do NOT silently
        # conclude None on empty evidence; flag it for resubmit instead.
        return {"accessible": True, "fetch_incomplete": True,
                "note": f"repo reachable but file-tree fetch failed (HTTP {st})"}
    tj = json.loads(tbody)
    paths = [x.get("path", "") for x in tj.get("tree", [])]
    truncated = tj.get("truncated", False)

    pipe_files = [p for p in paths if p.endswith(".pipe")]
    pipelines_folder = any(p == "pipelines" or "pipelines/" in p for p in paths)
    manifests = [p for p in paths if re.search(r"(^|/)(package\.json|requirements\.txt|pyproject\.toml)$", p)]
    claude_rule = any(p.endswith(".claude/rules/rocketride.md") for p in paths)

    dep_hits = []
    other = set()
    for mf in manifests[:8]:
        _, mb = _gh(f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{mf}")
        low = mb.lower()
        for line in mb.splitlines():
            if re.search(r"rocketride|@rocketride/sdk", line, re.I):
                dep_hits.append(f"{mf}: {line.strip()}")
        other.update(p for p in OTHER_PLATFORMS if p in low)

    # Broadened candidate set (includes api/route/config/server files), still bounded.
    kw = [p for p in paths if re.search(r"\.(ts|tsx|js|jsx|py|mjs)$", p)
          and re.search(r"rocket|pipeline|integration|agent|engine|config|api|route|server|main|index|app", p, re.I)]
    rest = [p for p in paths if re.search(r"\.(ts|tsx|js|jsx|py|mjs)$", p) and p not in kw]
    candidates = (kw + rest)[:22]
    source_signals = []
    for cf in candidates:
        _, cb = _gh(f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{cf}")
        low = cb.lower()
        for pat in SRC_PATS:
            if pat in low:
                source_signals.append(f"{cf}: {pat}")
        other.update(p for p in OTHER_PLATFORMS if p in low)

    has_real = bool(dep_hits or pipe_files or source_signals)
    return {
        "accessible": True,
        "default_branch": branch,
        "file_count": len(paths),
        "pipe_files": pipe_files[:12],
        "pipelines_folder": pipelines_folder,
        "rocketride_dependencies": dep_hits,
        "source_signals": source_signals[:20],
        "other_platforms": sorted(other),
        "claude_scaffold_only": claude_rule and not has_real,
        "tree_truncated": truncated,
    }


RUBRIC = (
    "You are a RocketRide hackathon usage verifier. Decide how a project used RocketRide "
    "using ONLY the code evidence provided below (already gathered from GitHub). Do not ask "
    "for more data and never invent details.\n\n"
    "REAL usage signals: a 'rocketride' or '@rocketride/sdk' dependency in a manifest; .pipe "
    "files or a pipelines/ folder; RocketRideClient / client.use() / client.chat() / .send(); "
    "live engine calls such as /v1/pipelines or ws://localhost:5565.\n"
    "NOT usage: a .claude/rules/rocketride.md file alone (default scaffold); branding or README "
    "mentions with no dependency or engine call.\n\n"
    "PRIMACY IS THE DECIDING QUESTION. First judge whether RocketRide is the primary orchestrator, "
    "then couple the tag to that.\n"
    "Backbone (pick one): 'Yes' = RocketRide runs the project's core AI / orchestration / pipeline "
    "logic — this is the PRIMARY backbone and the DEFAULT whenever RocketRide is the AI engine, EVEN "
    "IF another platform (Butterbase, Supabase, etc.) handles the database, model gateway, storage, "
    "or messaging. Handling data or the gateway does NOT demote RocketRide from 'Yes'. 'Partial' = "
    "ONLY when RocketRide is genuinely scoped to a single sub-feature while another platform runs the "
    "core app, OR there is a true co-equal split or a full non-RocketRide fallback path. 'No' = "
    "RocketRide is present but a minor/secondary component that does NOT run the core AI logic, OR it "
    "is only scaffold/branding, absent, or inaccessible.\n"
    "Tags (COUPLED to primacy): 'Significant' = real RocketRide integration (local .pipe files, OR "
    "HOSTED RocketRide pipelines called by id via the API / ROCKETRIDE_PIPELINE_ env + a rocketride "
    "adapter, OR RocketRideClient / live engine calls) AND RocketRide is the PRIMARY or CO-primary "
    "orchestrator (backbone Yes or Partial). 'Moderate' = real RocketRide artifacts exist (a .pipe or "
    "an integration) BUT RocketRide is NOT a primary orchestrator — another platform clearly runs the "
    "project (backbone No). 'Less' = RocketRide present only as branding/scaffold or explicitly "
    "disabled, no real artifacts. 'None' = no real usage, OR overclaimed (feedback claims RocketRide "
    "but the evidence shows none), OR inaccessible.\n"
    "DECISION RULE — apply strictly:\n"
    "1. No RocketRide signals at all -> 'None' / backbone 'No'. Only a lone bare mention, branding, or "
    "claude_scaffold_only=true -> 'Less' / backbone 'No'.\n"
    "2. Otherwise RocketRide is really used. Judge PRIMACY from the evidence + other_platforms: is "
    "RocketRide THE engine (backbone 'Yes'), a co-primary engine sharing with one other major platform "
    "(backbone 'Partial'), or present-but-not-primary while another platform runs the app (backbone "
    "'No')?\n"
    "3. Couple the tag to primacy: backbone 'Yes' or 'Partial' -> 'Significant'; backbone 'No' with "
    "real artifacts present -> 'Moderate'.\n"
    "4. A .pipe file ALONE does NOT make a project 'Significant': a real .pipe/integration earns at "
    "least 'Moderate', and reaches 'Significant' ONLY if RocketRide is also a primary or co-primary "
    "orchestrator. Hosted-pipeline usage (RocketRide pipelines called by id via the API) counts as "
    "real integration even with no local .pipe file.\n\n"
    "ARCHITECTURE LAYERS. Also decompose the project into five architectural layers and say what "
    "powers each, so a reviewer can see WHERE RocketRide sits:\n"
    " - ingest: how source data/content enters (upload, fetch, webhook, connectors).\n"
    " - retrieval: RAG / vector search / knowledge lookup (or 'none' if the project has none).\n"
    " - orchestration: what sequences the steps and runs the pipeline / agent logic.\n"
    " - reasoning: what makes the LLM / model calls (the AI engine).\n"
    " - output: how results are delivered (UI, API, file, message).\n"
    "For EACH layer emit exactly one of: 'rocketride' (RocketRide powers this layer), 'other' (a "
    "non-RocketRide tool powers it), or 'none' (the project has no such layer). orchestration and "
    "reasoning are the LOAD-BEARING layers and MUST stay consistent with backbone: backbone 'Yes' "
    "requires BOTH orchestration and reasoning = 'rocketride'; 'Partial' = exactly ONE of them "
    "'rocketride'; 'No' = neither.\n\n"
    "Output ONLY one strict JSON object (start with { and end with }), with these keys: "
    "repo_accessible (true), description (one plain-English sentence of what the project does), "
    "rocketride_usage (plain English: where RocketRide sits and what it does for the project — "
    "role and benefit, NOT file names), tag, backbone, justification (2-3 sentences explaining WHY "
    "this specific tag AND backbone were assigned — cite the concrete signals found and the primacy "
    "reasoning: why Significant vs Moderate/Less/None, and why Yes vs Partial vs No), "
    "layers (an object with EXACTLY the keys ingest, retrieval, orchestration, reasoning, output — "
    "each set to 'rocketride', 'other', or 'none' as defined above), notes (any "
    "extra caveats: overclaim, hidden usage, fallback, duplicate, etc.), evidence (array of short "
    "strings taken from the provided evidence). Keep description, rocketride_usage, and "
    "justification readable for a non-engineer judge."
)


# ---- pipeline driving --------------------------------------------------------

async def verify_one(client, token, row, sem) -> dict:
    project = row.get("project", "")
    url = row.get("github", "")
    feedback = row.get("feedback", "")
    started = time.perf_counter()

    if repo_missing(url):
        return {**row, "repo_accessible": False, "description": "", "rocketride_usage": "",
                "tag": "None", "backbone": "No",
                "notes": "No GitHub repo provided — flag for correction; scored as ZERO",
                "justification": "No GitHub repository was provided in the submission, so RocketRide "
                "usage cannot be verified from code — classified None / No and flagged for "
                "correction (score zero).",
                "evidence": [], "seconds": 0.0}

    # Python gathers the evidence off the event loop (keeps the websocket alive)
    sig = await asyncio.to_thread(fetch_signals, url)
    if not sig.get("accessible"):
        return {**row, "repo_accessible": False, "description": "", "rocketride_usage": "",
                "tag": "None", "backbone": "No",
                "notes": f"INACCESSIBLE (HTTP {sig.get('status', '?')}) — flag for correction: "
                         "double-check the repo URL; scored as ZERO",
                "justification": f"The repository could not be accessed (HTTP "
                f"{sig.get('status', '?')}), so RocketRide usage cannot be verified from code — "
                "classified None / No and flagged for correction (score zero).",
                "evidence": [], "seconds": round(time.perf_counter() - started, 1)}

    if sig.get("fetch_incomplete"):
        return {**row, "repo_accessible": None, "description": "", "rocketride_usage": "",
                "tag": "None", "backbone": "No",
                "notes": f"Evidence fetch incomplete ({sig.get('note', '')}) — resubmit this row",
                "justification": "Evidence gathering was incomplete this run, so classification was "
                "deferred — resubmit this row.",
                "evidence": [], "seconds": round(time.perf_counter() - started, 1)}

    prompt = (
        RUBRIC
        + f"\n\nPROJECT: {project}\nREPO: {url}\n"
        + f"TEAM FEEDBACK (may over/under-claim — trust the evidence over this): "
        + f"{feedback or '(none provided)'}\n\nCODE EVIDENCE (gathered from GitHub):\n"
        + json.dumps(sig, indent=2)
    )

    async with sem:
        parsed = {}
        for attempt in range(3):
            if attempt:
                await asyncio.sleep(1.0 * attempt)  # let classifier contention clear before retry
            q = Question()
            q.addQuestion(
                prompt if attempt == 0 else
                "Your previous reply was not valid JSON. Re-emit ONLY the strict JSON object "
                "for that project — start with { and end with }, no prose."
            )
            try:
                resp = await asyncio.wait_for(client.chat(token=token, question=q), timeout=60)
            except asyncio.TimeoutError:
                continue  # classifier hung on this call — retry, then flag if it never answers
            except Exception as e:  # noqa: BLE001
                return {**row, "repo_accessible": True, "classify_failed": True,
                        "tag": "None", "backbone": "No",
                        "notes": f"pipeline error: {e} — resubmit this row", "evidence": [],
                        "description": "", "rocketride_usage": "",
                        "seconds": round(time.perf_counter() - started, 1)}
            answers = resp.get("answers", []) if isinstance(resp, dict) else []
            parsed = extract_json(answers[0] if answers else "")
            if parsed:
                break

    elapsed = round(time.perf_counter() - started, 1)
    if not parsed:
        return {**row, "repo_accessible": True, "classify_failed": True,
                "tag": "None", "backbone": "No",
                "notes": "classifier returned no parseable JSON — resubmit this row",
                "evidence": [f"pipe_files={sig.get('pipe_files')}",
                             f"deps={sig.get('rocketride_dependencies')}"],
                "description": "", "rocketride_usage": "", "seconds": elapsed}
    parsed.setdefault("evidence", [])
    return {**row, **parsed, "seconds": elapsed}


async def run(rows, concurrency: int) -> list:
    sem = asyncio.Semaphore(concurrency)
    client = RocketRideClient()
    await client.connect()
    token = None
    try:
        result = await client.use(filepath=PIPELINE_FILE, use_existing=True)
        token = result["token"]
        print(f"Classifier live on Cloud (token: {token}); {len(rows)} rows, "
              f"concurrency={concurrency}, github_token={'yes' if GH_TOKEN else 'NO'}\n")
        tasks = [verify_one(client, token, r, sem) for r in rows]
        out = []
        for i, coro in enumerate(asyncio.as_completed(tasks), 1):
            res = await coro
            out.append(res)
            print(f"  [{i}/{len(rows)}] {res.get('project', '?'):42.42} "
                  f"-> {res.get('tag', '?'):11} / {res.get('backbone', '?'):7} "
                  f"({res.get('seconds', 0)}s)")
        return out
    finally:
        if token:
            try:
                await client.terminate(token)
            except Exception:
                pass
        await client.disconnect()


# ---- duplicate detection -----------------------------------------------------

def mark_duplicates(results: list) -> None:
    seen = {}
    for r in results:
        if repo_missing(r.get("github", "")):
            continue
        seen.setdefault(normalize_repo(r["github"]), []).append(r)
    for group in seen.values():
        if len(group) > 1:
            names = [g.get("project", "?") for g in group]
            for g in group:
                others = [n for n in names if n != g.get("project")]
                dup = f"duplicate repo (same as: {', '.join(others)})"
                g["notes"] = (g.get("notes", "") + "; " + dup).strip("; ")


# ---- sheet building ----------------------------------------------------------

TAG_FILL = {"Significant": "C6EFCE", "Moderate": "FFEB9C", "Less": "FCE4D6", "None": "FFC7CE"}
BACKBONE_FILL = {"Yes": "C6EFCE", "Partial": "FFEB9C", "No": "E7E6E6"}
WRAP = Alignment(wrap_text=True, vertical="top")

HEADERS = [
    "Project Name", "Team Details (Names / Emails)", "Project Description",
    "How RocketRide Was Used & How It Helped", "RocketRide Usage Tag",
    "RocketRide = Backbone?", "GitHub Link", "Additional Notes",
    "Why This Classification (Justification)",
    "Demo / Presentation / Video", "Deployed URL",
]


def row_values(r: dict) -> list:
    team = " / ".join(x for x in [r.get("names", ""), r.get("emails", "")] if x)
    notes = r.get("notes", "")
    evidence = r.get("evidence") or []
    if evidence:
        notes = (notes + " | evidence: " + "; ".join(map(str, evidence))).strip(" |")
    return [
        r.get("project", ""), team, r.get("description", ""), r.get("rocketride_usage", ""),
        r.get("tag", ""), r.get("backbone", ""), r.get("github", ""), notes,
        r.get("justification", ""),
        r.get("demo", ""), r.get("deployed", ""),
    ]


LINK_FONT = Font(color="0563C1", underline="single")


def _link_cell(ws, row_i: int, col: int) -> None:
    """Make a cell clickable, linking to the first URL it contains (handles bare github.com/www)."""
    cell = ws.cell(row=row_i, column=col)
    m = re.search(r"https?://\S+|(?:www\.|github\.com/)\S+", str(cell.value or ""), re.I)
    if m:
        url = m.group(0).rstrip(").,;")
        if not url.lower().startswith("http"):
            url = "https://" + url
        cell.hyperlink = url
        cell.font = LINK_FONT


def style_row(ws, row_i: int, r: dict) -> None:
    tag, bb = r.get("tag", ""), r.get("backbone", "")
    if tag in TAG_FILL:
        ws.cell(row=row_i, column=5).fill = PatternFill("solid", fgColor=TAG_FILL[tag])
    if bb in BACKBONE_FILL:
        ws.cell(row=row_i, column=6).fill = PatternFill("solid", fgColor=BACKBONE_FILL[bb])
    for col in (1, 2, 3, 4, 8, 9):
        ws.cell(row=row_i, column=col).alignment = WRAP
    for col in (7, 10, 11):  # GitHub Link, Demo/Presentation, Deployed URL — clickable
        _link_cell(ws, row_i, col)


def write_sheet(results: list, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Usage"
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = PatternFill("solid", fgColor="305496")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for r in results:
        ws.append(row_values(r))
        style_row(ws, ws.max_row, r)
    for i, w in enumerate([22, 30, 34, 50, 14, 14, 40, 40, 55, 38, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    wb.save(out_path)


def merge_into_sheet(results: list, existing_path: str, out_path: str) -> None:
    if not Path(existing_path).exists():
        sys.exit(f"--merge target not found: {existing_path} (run a full pass first)")
    wb = load_workbook(existing_path)
    ws = wb.active

    def _key(s):
        return re.sub(r"\s+", " ", str(s or "").strip().lower())

    name_to_row, repo_to_row = {}, {}
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(row=r, column=1).value
        if nm:
            name_to_row[_key(nm)] = r
        gh = ws.cell(row=r, column=7).value or ""
        if "github.com" in gh.lower():
            repo_to_row[normalize_repo(gh)] = r

    updated, appended = [], []
    for res in results:
        gh = res.get("github", "")
        row_i = name_to_row.get(_key(res.get("project", "")))
        if not row_i and "github.com" in gh.lower():
            row_i = repo_to_row.get(normalize_repo(gh))   # fall back to repo match
        if row_i:                                          # UPDATE existing
            for col, val in enumerate(row_values(res), 1):
                ws.cell(row=row_i, column=col).value = val
            style_row(ws, row_i, res)
            updated.append(res.get("project", ""))
        else:                                              # APPEND new
            ws.append(row_values(res))
            style_row(ws, ws.max_row, res)
            appended.append(res.get("project", ""))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    wb.save(out_path)
    print(f"\nUpdated {len(updated)} existing row(s); appended {len(appended)} new row(s) -> {out_path}")
    if appended:
        print("  appended:", ", ".join(appended))


def write_needs_review(failed: list, path: str) -> None:
    cols = ["Project title", "Team Members Names", "Team Members Emails", "Github Repo", "Feedback"]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for r in failed:
            w.writerow([r.get("project", ""), r.get("names", ""), r.get("emails", ""),
                        r.get("github", ""), r.get("feedback", "")])


def print_summary(results: list) -> None:
    from collections import Counter
    tags = Counter(r.get("tag", "?") for r in results)
    bb = Counter(r.get("backbone", "?") for r in results)
    times = [r.get("seconds", 0) for r in results if r.get("seconds")]
    print("\n=== Summary ===")
    print("Tags:    ", dict(tags))
    print("Backbone:", dict(bb))
    if times:
        print(f"Per-repo: avg {sum(times)/len(times):.1f}s, max {max(times):.1f}s (n={len(times)})")


# ---- main --------------------------------------------------------------------

def main() -> None:
    global GH_TOKEN
    ap = argparse.ArgumentParser()
    ap.add_argument("input", help="submissions .csv/.xlsx (or the corrected needs_review.csv)")
    ap.add_argument("--merge", help="existing sheet to PATCH re-verified rows into")
    ap.add_argument("--out", help="output path (default: new sheet, or the --merge file)")
    ap.add_argument("--concurrency", type=int, default=2)
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()

    GH_TOKEN = github_token()
    rows = load_rows(args.input)
    if args.limit:
        rows = rows[: args.limit]
    print(f"Loaded {len(rows)} rows from {args.input}"
          + (f"  (MERGE into {args.merge})" if args.merge else ""))

    t0 = time.perf_counter()
    results = asyncio.run(run(rows, args.concurrency))
    print(f"\nTotal batch wall-clock: {time.perf_counter() - t0:.0f}s for {len(rows)} rows "
          f"(this is the real SLA number; the per-row seconds above are cumulative)")
    mark_duplicates(results)

    if args.merge:
        merge_into_sheet(results, args.merge, args.out or args.merge)
    else:
        out = args.out or "RocketRide_Hackathon_Usage.xlsx"
        write_sheet(results, out)
        print(f"\nWrote {len(results)} rows -> {out}")
        failed = [r for r in results if is_failed(r)]
        if failed:
            write_needs_review(failed, NEEDS_REVIEW)
            print(f"{len(failed)} row(s) failed the access check -> {NEEDS_REVIEW}. "
                  f"Fix the URLs there, then:\n"
                  f'  python run_batch.py "{NEEDS_REVIEW}" --merge "{out}"')

    print_summary(results)


if __name__ == "__main__":
    main()
